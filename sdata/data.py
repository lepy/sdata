# -*-coding: utf-8-*-
from __future__ import division

'''
basic sdata types 
'''
from sdata import __version__
import sys
import os
import uuid
from collections import OrderedDict
import logging
logger = logging.getLogger("sdata")
import numpy as np
import pandas as pd
import shutil
import copy
from sdata.metadata import Metadata, Attribute, extract_name_unit
from sdata.timestamp import now_utc_str, now_local_str, today_str
import inspect
import json
import hashlib
import base64
import requests
from tabulate import tabulate

if sys.version_info < (3, 0):
    from StringIO import StringIO
else:
    from io import BytesIO, StringIO

class Sdata_Name_Exeption(Exception): pass
class Sdata_Uuid_Exeption(Exception): pass

if sys.version_info < (3, 6):
    import sha3

try:
    import openpyxl
except:
    logger.warning("openpyxl is not available -> no xlsx import")

def uuid_from_str(name):
    return uuid.uuid3(uuid.NAMESPACE_DNS, name)

class Data(object):
    """Base sdata object"""
    ATTR_NAMES = []

    SDATA_VERSION = "!sdata_version"
    SDATA_NAME = "!sdata_name"
    SDATA_UUID = "!sdata_uuid"
    SDATA_CTIME = "!sdata_ctime"
    SDATA_MTIME = "!sdata_mtime"
    SDATA_PARENT = "!sdata_parent"
    SDATA_CLASS = "!sdata_class"
    SDATA_PROJECT = "!sdata_project"

    def __init__(self, **kwargs):
        """create Data object

        .. code-block:: python

            df = pd.DataFrame([1,2,3])
            data = sdata.Data(name='my name',
                        uuid='38b26864e7794f5182d38459bab85842',
                        table=df,
                        description="A remarkable description")


        :param name: name of the data object
        :param table: pandas.DataFrame to store
        :param uuid: uuid of the object
        :param metadata: sdata.Metadata object
        :param description: a string to describe the object
        """

        # self._uuid = None
        # self._name = None
        self._prefix = None

        # ToDo: add getter and setter for metadata
        # self.metadata = kwargs.get("metadata") or Metadata()

        self.metadata = Metadata()

        # set default sdata attributes
        self.metadata.add(self.SDATA_VERSION, __version__, dtype="str", description="sdata package version")
        self.metadata.add(self.SDATA_NAME, "N.N.", dtype="str", description="name of the data object")
        self.metadata.add(self.SDATA_UUID, "", dtype="str", description="Universally Unique Identifier")
        self.metadata.add(self.SDATA_PARENT, "", dtype="str", description="uuid of the parent sdata object")
        self.metadata.add(self.SDATA_CLASS, self.__class__.__name__, dtype="str", description="sdata class")
        self.metadata.add(self.SDATA_CTIME, now_utc_str(), dtype="str", description="creation date")
        self.metadata.add(self.SDATA_MTIME, now_utc_str(), dtype="str", description="modification date")

        metadata = kwargs.get("metadata")
        if metadata is not None:
            # logger.debug("Data got Metadata {}".format(metadata))
            if metadata and isinstance(metadata, Metadata):
                for attribute in metadata.attributes.values():
                    # logger.debug("Data.Metadata.add {0.name}:{0.value}".format(attribute))
                    self.metadata.add(attribute)

        # auto correct
        if kwargs.get("auto_correct") is None or kwargs.get("auto_correct") is True:
            self.auto_correct = True
        else:
            self.auto_correct = False
        # logger.debug("sdata: set auto_correct={}".format(self.auto_correct))

        if kwargs.get("name") is not None:
            self.name = kwargs.get("name")

        self.prefix = kwargs.get("prefix") or ""
        self._gen_default_attributes(kwargs.get("default_attributes") or self.ATTR_NAMES)
        self._group = OrderedDict()
        self._table = None  # pd.DataFrame()
        self.table = kwargs.get("table", None)
        self._description = ""
        self.description = kwargs.get("description", "")
        self.project = kwargs.get("project", "")

        if (kwargs.get("uuid")=="" or kwargs.get("uuid") is not None) and not self.metadata.get(self.SDATA_UUID).value and kwargs.get("uuid")!="hash":
            # logger.info("uuid in kwargs")
            try:
                self._set_uuid(kwargs.get("uuid")) # store given uuid str or generate a new uuid
            except Sdata_Uuid_Exeption as exp:
                if self.auto_correct is True:
                    logger.warning("got invalid uuid -> generate a new uuid")
                    self._set_uuid(uuid.uuid4())
                else:
                    raise
        elif (kwargs.get("uuid")=="" or kwargs.get("uuid") is None) and self.metadata.get(self.SDATA_UUID).value != "":
            # logger.info("uuid in metadata")
            pass
        elif kwargs.get("uuid")=="hash":
            sha3_256 = self.gen_uuid_from_state()
            # logger.info("gen uuid from sha3_256 {}".format(sha3_256))
            new_uuid = uuid_from_str(sha3_256)
            self._set_uuid(new_uuid.hex)
        else:
            # logger.info("uuid new")
            self._set_uuid(uuid.uuid4())

    def gen_uuid_from_state(self):
        """generate the same uuid for the same data

        :return: uuid
        """
        s = hashlib.sha3_256()
        metadata = self.metadata.copy()
        metadata.attributes.pop(self.SDATA_UUID)
        metadata.attributes.pop(self.SDATA_MTIME)
        metadata.attributes.pop(self.SDATA_CTIME)
        metadatastr = metadata.to_json().encode(errors="replace")
        s.update(metadatastr)
        if self.table is not None:
            tablestr = self.table.to_json().encode(errors="replace")
            s.update(tablestr)
        s.update(self.description.encode(errors="replace"))
        return s.hexdigest()

    def __eq__(self, other):
        """compare Data checksum
        
        :param other: sdata.Data objecet
        :return: True or False
        """
        if not isinstance(other, self.__class__):
            logger.debug("you should not compare {} with {}!".format(self.__class__.__name__, other.__class__.__name__))
            return False
        return self.sha3_256 == other.sha3_256

    def update_mtime(self):
        """update modification time

        :return:
        """
        self.metadata.add(self.SDATA_MTIME, now_utc_str())

    @property
    def sha3_256_table(self):
        """Return a SHA3 hash of the sData.table object with a hashbit length of 32 bytes.

        .. code-block:: python

            sdata.Data(name="1", uuid=sdata.uuid_from_str("1")).sha3_256_table

            'c468e659891eb5dea6eb6baf73f51ca0688792bf9ad723209dc22730903f6efa'

        :return: hashlib.sha3_256.hexdigest()
        """
        s = hashlib.sha3_256()
        if self.table is not None:
            tablestr = self.table.to_json().encode(errors="replace")
            s.update(tablestr)
        return s.hexdigest()

    @property
    def sha3_256(self):
        """Return a SHA3 hash of the sData object with a hashbit length of 32 bytes.

        .. code-block:: python

            sdata.Data(name="1", uuid=sdata.uuid_from_str("1")).sha3_256

            'c468e659891eb5dea6eb6baf73f51ca0688792bf9ad723209dc22730903f6efa'

        :return: hashlib.sha3_256.hexdigest()
        """
        s = hashlib.sha3_256()
        metadatastr = self.metadata.to_json().encode(errors="replace")
        s.update(metadatastr)
        if self.table is not None:
            tablestr = self.table.to_json().encode(errors="replace")
            s.update(tablestr)
        s.update(self.description.encode(errors="replace"))
        return s.hexdigest()

    def update_hash(self, hashobject):
        """A hash represents the object used to calculate a checksum of a
        string of information.

        .. code-block:: python

            data = sdata.Data()

            md5 = hashlib.md5()
            data.update_hash(md5)
            md5.hexdigest()
            'bbf323bdcb0bf961803b5504a8a60d69'

            sha1 = hashlib.sha1()
            data.update_hash(sha1)
            sha1.hexdigest()
            '3c59368c7735c1ecaf03ebd4c595bb6e73e90f0c'

            hashobject = hashlib.sha3_256()
            data.update_hash(hashobject).hexdigest()
            'c468e659891eb5dea6eb6baf73f51ca0688792bf9ad723209dc22730903f6efa'

            data.update_hash(hashobject).digest()
            b'M8...'

        :param hash: hash object, e.g. hashlib.sha1()
        :return: hash
        """
        if not (hasattr(hashobject, "update") and hasattr(hashobject, "hexdigest")):
            logger.error("Data.update_hash: given hashfunction is invalid")
            raise Exception("Data.update_hash: given hashfunction is invalid")

        metadatastr = self.metadata.to_json().encode(errors="replace")
        hashobject.update(metadatastr)
        if self.table is not None:
            tablestr = self.table.to_json().encode(errors="replace")
            hashobject.update(tablestr)
        hashobject.update(self.description.encode(errors="replace"))
        return hashobject

    def describe(self):
        """Generate descriptive info of the data

        .. code-block:: python

            df = pd.DataFrame([1,2,3])
            data = sdata.Data(name='my name',
                        uuid='38b26864e7794f5182d38459bab85842',
                        table=df,
                        description="A remarkable description")
            data.describe()

        .. code-block:: none

                            0
            metadata        3
            table_rows      3
            table_columns   1
            description    24


        :return: pd.DataFrame
        """
        df = pd.DataFrame({0: []}, dtype=object)
        df.loc["metadata", 0] = self.metadata.size
        if self.table is None:
            df.loc["table_rows"] = 0
            df.loc["table_columns"] = 0
        else:
            df.loc["table_rows"] = len(self.table)
            df.loc["table_columns"] = len(self.table.columns)
        df.loc["description", 0] = len(self.description)
        return df

    def _gen_default_attributes(self, default_attributes):
        """create default Attributes in data.metadata"""
        for attr_name, value, dtype, unit, description, required in default_attributes:
            self.metadata.set_attr(name=attr_name, value=value, dtype=dtype, description=description)

    def _get_uuid(self):
        return self.metadata.get(self.SDATA_UUID).value
        # return self._uuid

    def _set_uuid(self, value):
        if isinstance(value, str):
            try:
                uuid.UUID(value)
                self.metadata.set_attr(self.SDATA_UUID, uuid.UUID(value).hex)
            except ValueError as exp:
                logger.warning("data.uuid: %s" % exp)
                raise Sdata_Uuid_Exeption("got invalid uuid str '{}'".format(str(value)))
        elif isinstance(value, uuid.UUID):
            self.metadata.set_attr(self.SDATA_UUID, value.hex)
        else:
            logger.error("Data.uuid: invalid uuid '{}'".format(value))
            raise Exception("Data.uuid: invalid uuid '{}'".format(value))

    uuid = property(fget=_get_uuid, fset=_set_uuid, doc="uuid of the object")

    def _get_name(self):
        # return self._name
        return self.metadata.get(self.SDATA_NAME).value

    def _set_name(self, value):
        if isinstance(value, str):
            try:
                self.metadata.set_attr(self.SDATA_NAME, str(value)[:256])
            except ValueError as exp:
                logger.warning("data.name: %s" % exp)
        else:
            # self._name = str(value)[:256]
            self.metadata.set_attr(self.SDATA_NAME, str(value)[:256])

    name = property(fget=_get_name, fset=_set_name, doc="name of the object")

    def _get_project(self):
        return self.metadata.get(self.SDATA_PROJECT).value

    def _set_project(self, value):
        if isinstance(value, str):
            try:
                self.metadata.set_attr(self.SDATA_PROJECT, str(value)[:256])
            except ValueError as exp:
                logger.warning("data.project: %s" % exp)
        else:
            # self._name = str(value)[:256]
            self.metadata.set_attr(self.SDATA_PROJECT, str(value)[:256])

    project = property(fget=_get_project, fset=_set_project, doc="name of the project")

    def _get_description(self):
        return self._description

    def _set_description(self, value):
        if isinstance(value, str):
            try:
                self._description = str(value)
            except ValueError as exp:
                logger.warning("data.name: %s" % exp)
        else:
            self._description = str(value)

    description = property(fget=_get_description, fset=_set_description, doc="description of the object")

    @property
    def filename(self):

        validchars = "-_.() "
        out = ""

        name = "{}".format(self.name)

        for c in name:
            if str.isalpha(c) or str.isdigit(c) or (c in validchars):
                out += c
            else:
                out += "_"
        return out

    def _get_prefix(self):
        return self._prefix

    def _set_prefix(self, value):
        if isinstance(value, str):
            try:
                self._prefix = value[:256]
            except ValueError as exp:
                logger.warning("data.prefix: %s" % exp)
        else:
            self._prefix = str(value)[:256]

    prefix = property(fget=_get_prefix, fset=_set_prefix, doc="prefix of the object name")

    def _get_table(self):
        return self._table

    def _set_table(self, df):
        if isinstance(df, pd.DataFrame):
            self._table = df
            self._table.index.name = "index"

    table = property(fget=_get_table, fset=_set_table, doc="table object(pandas.DataFrame)")
    df = table

    def description_to_df(self):
        """get description as DataFrame

        :return: DataFrame of description lines
        """
        return pd.DataFrame(self.description.splitlines())

    def description_from_df(self, df):
        """set description from DataFrame of lines

        :return:
        """
        if df is not None and isinstance(df, pd.DataFrame) and len(df)>0:
            lines = df.iloc[:, 0]
            lines = lines.astype(str)
            self.description = "\n".join(lines.values)

    def to_folder(self, path, dtype="csv"):
        """export data to folder

        :param path:
        :param dtype:
        :return:
        """

        if dtype not in ["csv", "xlsx"]:
            dtype = "xlsx"
        if not os.path.exists(path):
            try:
                os.makedirs(path)
            except OSError as exp:
                logger.error(exp)
        else:
            self.clear_folder(path)

        self.metadata.set_attr(name="class", value=self.__class__.__name__, description="object class", unit="-",
                               dtype="str")
        self.metadata.set_attr(name="uuid", value=self.uuid, description="object uuid", unit="-", dtype="str")
        self.metadata.set_attr(name="name", value=self.name, description="object name", unit="-", dtype="str")

        if dtype == "csv":
            metadata_filepath = os.path.join(path, "metadata.csv")
            logger.debug("export meta csv '{}'".format(metadata_filepath))
            self.metadata.to_csv(metadata_filepath)

            # table export
            if isinstance(self._table, pd.DataFrame) and len(self._table) > 0:
                exportpath = os.path.join(path, "{}.csv".format(self.osname))
                self._table.to_csv(exportpath, index=False)
        if dtype == "xlsx":
            if not isinstance(self._table, pd.DataFrame):
                self.table = pd.DataFrame()
            exportpath = os.path.join(path, "{}.xlsx".format(self.osname))
            self.to_xlsx(exportpath)
        # group export
        for data in self.group.values():
            exportpath = os.path.join(path, "{}-{}".format(data.__class__.__name__.lower(), data.osname))
            data.to_folder(exportpath, dtype=dtype)
        return path

    @classmethod
    def from_folder(cls, path):
        """sdata object instance

        :param path:
        :return:
        """
        # data = Data.from_folder(path)

        data = cls()
        if not os.path.exists(path):
            logger.error("from_folder error: path '{}' not exists.".format(path))
            return data

        data.metadata = data._load_metadata(path)
        try:
            data.uuid = data.metadata.get_attr("uuid").value
            data.name = data.metadata.get_attr("name").value
        except Exception as exp:
            logger.error("Data.from_folder: {}".format(data.metadata.to_dict()))
            raise

        # table import
        files = [x for x in os.listdir(path) if
                 not os.path.isdir(os.path.join(path, x)) and not x.startswith("metadata")]
        if len(files) == 1:
            assert len(files) == 1, "invalid number of files for Table '{}'".format(files)
            importpath = os.path.join(path, files[0])
            print("read table {}".format(importpath))
            # data._table = pd.read_csv(importpath)

        if not os.path.exists(path):
            return cls()
        metadata = cls._load_metadata(path)
        data = cls()
        data.metadata = metadata
        data.uuid = data.metadata.get_attr("uuid").value
        data.name = data.metadata.get_attr("name").value

        folders = [x for x in os.listdir(path) if os.path.isdir(os.path.join(path, x))]
        for folder in folders:
            subfolder = os.path.join(path, folder)
            data_ = data.from_folder(subfolder)
            subdata = data_.from_folder(subfolder)
            data.add_data(subdata)
        return data

    @staticmethod
    def clear_folder(path):
        """delete subfolder in export folder

        :param path: path
        :return: None
        """

        def is_valid(path):
            prefix = path.split("-")[0]
            if prefix in [x.lower() for x in SDATACLS.keys()]:
                return True
            else:
                return False

        subfolders = [x for x in os.listdir(path) if os.path.isdir(os.path.join(path, x))]
        valid_subfolders = [x for x in subfolders if is_valid(x)]
        for subfolder in valid_subfolders:
            try:
                subfolder = os.path.join(path, subfolder)
                logger.debug("clear_folder: rm {}".format(subfolder))
                shutil.rmtree(subfolder)
            except OSError as exp:
                raise

    @staticmethod
    def _load_metadata(path):
        """load metadata from csv

        :returns: Metadata instance"""
        metadata_filepath = os.path.join(path, "metadata.csv")
        if os.path.exists(metadata_filepath):
            metadata = Metadata().from_csv(metadata_filepath)
        else:
            metadata = Metadata()
        return metadata

    @staticmethod
    def _get_class_from_metadata(metadata):
        """get class object from metadata

        :returns: relevant sdata class object"""
        classattr = metadata.get_attr("class")
        if classattr is not None:
            sdataclassname = classattr.value
            sdatacls = SDATACLS.get(sdataclassname)
            if sdataclassname not in SDATACLS:
                logger.warning("unsupported cls '{}'".format(sdataclassname))
                sdatacls = Data
        else:
            logger.warning("cls not defined '{}'".format(metadata))
            sdatacls = None
        return sdatacls

    @property
    def osname(self):
        """:returns: os compatible name (ascii?)"""
        return self.asciiname.lower()

    @property
    def asciiname(self):
        name = copy.copy(self.name)
        mapper = [("ä", "ae"), ("ö", "oe"), ("ü", "ue"), ("Ä", "Ae"), ("Ö", "Oe"), ("Ü", "Ue"),
                  ("ß", "sz"), (" ", "_"), ("/", "_"), ("\\", "_")]
        for k, v in mapper:
            name = name.replace(k, v)
        return name.encode('ascii', 'replace').decode("ascii")

    def verify_attributes(self):
        """check mandatory attributes"""
        invalid_attrs = []
        # attr_defs = ["name", "value", "dtype", "unit", "description", "required"]
        for attr_defs in self.ATTR_NAMES:
            required = attr_defs[5]
            if required is False:
                continue
            attr = self.metadata.get_attr(attr_defs[0])
            if attr is None:
                invalid_attrs.append(attr_defs[0])
            elif attr.value is None:
                invalid_attrs.append(attr.name)
        return invalid_attrs

    def __str__(self):
        return "(Data '%s':%s)" % (self.name, self.uuid)

    __repr__ = __str__

    def get_group(self):
        return self._group

    group = property(get_group, doc="get group")

    def keys(self):
        """get all child objects uuids

        :return: list of uuid's
        """
        return list(self.group.keys())

    def values(self):
        """get all child objects

        :return: list of child objects
        """
        return list(self.group.values())

    def items(self):
        """get all child objects

        :return: [(child uuid, child objects), ]
        """
        return list(self.group.items())

    def clear_group(self):
        """clear group dict"""
        self._group = OrderedDict()

    def add_data(self, data):
        """add data, if data.name is unique"""
        if isinstance(data, Data):
            names = [dat.name.lower() for uid, dat in self.group.items()]
            if data.name.lower() in names:
                logger.error("{}: name '{}' aready exists".format(data.__class__.__name__, data.name))
                return
            self.group[data.uuid] = data
        else:
            logger.warning("ignore data %s (wrong type!)" % data)

    def get_data_by_uuid(self, uid):
        """get data by uuid"""
        return self.group.get(uid)

    def get_data_by_name(self, name):
        """:return obj by name"""
        d = dict([(obj.name, uid) for uid, obj in self.group.items()])
        uid = d.get(name)
        return self.get_data_by_uuid(uid)

    def tree_folder(self, dir, padding="  ", print_files=True, hidden_files=False, last=True):
        """print tree folder structure"""
        if last is False:
            print(padding[:-1] + '├─' + os.path.basename(os.path.abspath(dir)))
        else:
            print(padding[:-1] + '└─' + os.path.basename(os.path.abspath(dir)))
        padding = padding + ' '
        files = []
        if print_files:
            files = [x for x in sorted(os.listdir(dir)) if not x.startswith(".")]
        else:
            files = [x for x in sorted(os.listdir(dir)) if os.path.isdir(dir + os.sep + x)]

        # metadata first
        metafiles = [f for f in files if f.startswith("metadata")]
        files = [x for x in files if x not in metafiles]
        files = metafiles + sorted(files)

        for count, file in enumerate(sorted(files)):
            # print(padding + '|')
            path = dir + os.sep + file
            if os.path.isdir(path):
                if count == (len(files) - 1):
                    self.tree_folder(path, padding + ' ', print_files, last=True)
                else:
                    self.tree_folder(path, padding + '|', print_files, last=False)
            else:
                if count == (len(files) - 1):
                    print(padding + '└─' + file)
                else:
                    print(padding + '├─' + file)

    def dir(self):
        """returns a nested list of all child objects

        :return: list of sdata.Data objects
        """
        return [(x.name, x.dir()) for x in self.group.values()]

    def to_xlsx_byteio(self):
        """get xlsx as byteio

        :return: BytesIO
        """

        def adjust_col_width(sheetname, df, writer, width=40):
            worksheet = writer.sheets[sheetname]  # pull worksheet object
            worksheet.set_column(0, 0, width)
            for idx, col in enumerate(df):  # loop through all columns
                # series = df[col]
                # max_len = max((
                #     series.astype(str, raise_on_error=False).map(len).max(),  # len of largest item
                #     len(str(series.name))  # len of column name/header
                #     )) + 1  # adding a little extra space
                worksheet.set_column(idx + 1, idx + 1, width)

        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        self.metadata.df.to_excel(writer, sheet_name='metadata')
        adjust_col_width('metadata', self.metadata.df, writer)

        self.df.to_excel(writer, sheet_name='table')
        adjust_col_width('table', self.table, writer, width=15)

        df_description = pd.DataFrame(self.description.splitlines())
        df_description.to_excel(writer, sheet_name='description', index=False, header=None)
        adjust_col_width('description', df_description, writer, width=200)

        writer.save()
        processed_data = output.getvalue()
        return processed_data

    def to_xlsx_base64(self):
        """get xlsx as byteio base64 encoded

        :return: base64
        """
        val = self.to_xlsx_byteio()
        b64 = base64.b64encode(val)
        return b64

    def get_download_link(self):
        """Generates a link allowing the data in a given panda dataframe to be downloaded
        in:  dataframe
        out: href string
        """
        b64 = self.to_xlsx_base64()
        return '<a href="data:application/octet-stream;base64,{1}" download="{0}.xlsx">Download {0}.xlsx file</a>'.format(
            self.osname, b64.decode())

    def to_xlsx(self, filepath=None):
        """export atrributes and data to excel

        :param filepath:
        :return:
        """

        def adjust_col_width(sheetname, df, writer, width=40):
            worksheet = writer.sheets[sheetname]  # pull worksheet object
            worksheet.set_column(0, 0, width)
            for idx, col in enumerate(df):  # loop through all columns
                # series = df[col]
                # max_len = max((
                #     series.astype(str, raise_on_error=False).map(len).max(),  # len of largest item
                #     len(str(series.name))  # len of column name/header
                #     )) + 1  # adding a little extra space
                worksheet.set_column(idx + 1, idx + 1, width)

        with pd.ExcelWriter(filepath) as writer:

            # metadata
            # dfm = pd.DataFrame.from_dict(self.metadata, orient="index", columns=["value"])
            dfm = self.metadata.to_dataframe()

            # dfm = dfm.sort_index()
            dfm.index.name = "key"
            dfm.to_excel(writer, sheet_name='metadata', index=False)
            adjust_col_width('metadata', dfm, writer)

            # data
            if self.table is not None:
                self.table.index.name = "index"
                self.table.to_excel(writer, sheet_name='table')
                adjust_col_width('table', self.table, writer, width=15)
            else:
                df = pd.DataFrame()
                df.index.name = "index"
                df.to_excel(writer, sheet_name='table')
                adjust_col_width('table', df, writer, width=15)

            df_description = pd.DataFrame(self.description.splitlines())
            df_description.to_excel(writer, sheet_name='description', index=False, header=None)
            adjust_col_width('description', df_description, writer, width=200)

            # # raw data
            # self.df_raw.index.name = "index"
            # self.df_raw.to_excel(writer, sheet_name='df_raw')
            # adjust_col_width('df_raw', self.df_raw, writer, width=15)

    @classmethod
    def from_xlsx(cls, filepath):
        """save table as xlsx

        :param filepath:
        :return:
        """
        try:
            if os.path.exists(filepath):
                wb = openpyxl.load_workbook(filename=filepath)
                # sheetname = u'Übergabedaten VWD für BFA'
                sheetnames = wb.sheetnames

                tt = cls(name=filepath)

                # read df
                if "table" in sheetnames:
                    tt.table = pd.read_excel(filepath, sheet_name="table", index_col='index')
                else:
                    logger.info("no table data in '{}'".format(filepath))
                dfm = pd.read_excel(filepath, sheet_name="metadata")
                dfm = dfm.set_index(dfm.name.values)
                # dfm["value"] = dfm["value"].replace(np.nan, None)
                dfm["description"] = dfm["description"].replace(np.nan, '')
                dfm["label"] = dfm["label"].replace(np.nan, '')
                # print("!data.from_xlsx", dfm)
                tt.metadata = tt.metadata.from_dataframe(dfm)

                # read description
                if "description" in sheetnames:
                    cells = []
                    for cell in wb["description"]["A"]:
                        if cell.value is not None:
                            cells.append(cell.value)
                        else:
                            cells.append("")
                    tt.description = "\n".join(cells)
                else:
                    logger.info("no description in '{}'".format(filepath))

                return tt
            else:
                raise Exception("excel file '{}' not available".format(filepath))
        except Exception as exp:
            raise

    def to_json(self, filepath=None):
        """export Data in json format

        :param filepath: export file path (default:None)
        :return: json str
        """

        if self.table is not None:
            json_table = self.table.to_dict()
        else:
            json_table = {}

        j = {"metadata": self.metadata.to_dict(),
             "table": json_table,
             "description": self.description
             }
        if filepath:
            with open(filepath, "w") as fh:
                json.dump(j, fh)
        else:
            return json.dumps(j)

    @classmethod
    def from_json(cls, s=None, filepath=None):
        """create Data from json str or file

        :param s: json str
        :param filepath:
        :return: sdata.Data
        """
        data = cls(name="N.N.")
        if s is None and filepath is not None:
            with open(filepath, "r") as fh:
                d = json.load(fh)
        elif s is None and filepath is None:
            logger.error("data.from_json: no json data available")
            return
        elif s is not None and filepath is None:
            d = json.loads(s)
        else:
            logger.error("data.from_json: unexpected error")
            d = None

        if d:
            if "metadata" in d.keys():
                data.metadata.update_from_dict(d["metadata"])
            else:
                logger.error("Data.from_json: table not available")

            if "table" in d.keys():
                data.table = pd.DataFrame.from_dict(d["table"])
                # data.table = pd.read_json(json.dumps(d["table"]))
                # data.table = pd.read_json(d["table"])
            else:
                logger.error("Data.from_json: metadata not available")

            if "description" in d.keys():
                data.description = d["description"]
            else:
                logger.error("Data.from_json: description not available")

        return data

    @classmethod
    def from_url(cls, url=None, stype=None):
        """create Data from json str or file

        :param url: url
        :param stype: "json" ("xlsx", "csv")
        :return: sdata.Data
        """

        supported_stypes = ["json"]

        if stype not in supported_stypes:
            raise NotADirectoryError("stype '{}' is not supported".format(stype))
            return

        raw = requests.get(url).text
        if  stype=="json":
            data = cls.from_json(raw)
            return data


    def to_csv(self, filepath=None):
        """export sdata.Data to csv

        :param filepath:
        :return:
        """
        exportlines = []
        exportlines.append(self.metadata.to_csv_header(prefix="#;", sep=";", filepath=None))
        exportlines.append(self.df.to_csv(sep=";"))

        exportstr = "".join(exportlines)

        if filepath is None:
            return exportstr
        else:
            with open(filepath, "w") as fh:
                fh.write(exportstr)

    @classmethod
    def from_csv(cls, s=None, filepath=None, sep=";"):
        """import sdata.Data from csv

        :param s: csv str
        :param filepath:
        :param sep: separator (default=";")
        :return: sdata.Data
        """
        data = cls()
        if filepath:
            df = pd.read_csv(filepath, sep=";", comment="#")
            sio = open(filepath, "r")
        elif s is not None:
            sio = StringIO(s)
            pd.read_csv(sio, sep=";", comment="#")
            sio.seek(0)
        else:
            logger.error("data.from_csv: no csv data available")
            return

        for line in sio:
            if line.startswith("#;"):
                print(line)

        data.table = df
        return data

    def to_hdf5(self, filepath, **kwargs):
        """export sdata.Data to hdf5

        :param filepath:
        :param complib: default='zlib' ['zlib', 'lzo', 'bzip2', 'blosc', 'blosc:blosclz', 'blosc:lz4', 'blosc:lz4hc', 'blosc:snappy', 'blosc:zlib', 'blosc:zstd']
        :param complevel: default=9 [0-9]

        :return:
        """
        if not isinstance(self.df, pd.DataFrame):
            df = pd.DataFrame()
        else:
            df = self.df
        kwargs["mode"] = "w"
        if kwargs.get("complib") is None:
            kwargs["complib"] = "zlib"

        if kwargs.get("complevel") is None:
            kwargs["complevel"] = 9

        with pd.HDFStore(filepath, **kwargs) as hdf:
            hdf.put('metadata'.format(self.uuid), self.metadata.df, format='fixed', data_columns=True)
            hdf.put('table'.format(self.uuid), df, format='fixed', data_columns=True)
            hdf.put('description'.format(self.uuid), self.description_to_df(), format='fixed', data_columns=True)


    @classmethod
    def from_hdf5(cls, filepath, **kwargs):
        """import sdata.Data from hdf5

        :param filepath:
        :param sep: separator (default=";")
        :return: sdata.Data
        """
        if not os.path.exists:
            logger.error("hdf5 file '' notr available".format(filepath))
            return
        
        with pd.HDFStore(filepath, mode="r+") as hdf:
            metadata_path = "/metadata".format(uuid)
            table_path = "/table".format(uuid)
            description_path = "/description".format(uuid)
            df_metadata = hdf.get(metadata_path)
            df_table = hdf.get(table_path)
            df_description = hdf.get(description_path)
            metadata = Metadata.from_dataframe(df_metadata)
            # logger.debug("hdf {}".format(metadata.get("!sdata_uuid").value))
            data = Data(metadata=metadata, table=df_table)
            data.description_from_df(df_description)

        return data

    def to_html(self, filepath, xlsx=True, style=None):
        """export Data to html

        :param filepath:
        :param xlsx:
        :param style:
        :return:
        """

        table_values = self.df.head()
        table_headers = self.df.columns

        table_description_values = self.df.describe()
        table_description_headers = self.df.describe().columns

        metadata_values = self.metadata.df.head().values
        metadata_headers = self.metadata.df.columns

        if xlsx is True:
            xlsx_tag = self.get_download_link()
        else:
            xlsx_tag = ""

        param = {"title":"{0} [{1}]".format(self.osname,
                                            self.uuid),
                 "description":self.description,
                 "metadata": tabulate(metadata_values, metadata_headers, tablefmt="html"),
                 "table": tabulate(table_values, table_headers, tablefmt="html"),
                 "table_description": tabulate(table_description_values, table_description_headers, tablefmt="html"),
                 "xlsx_tag":xlsx_tag,
                 "sdata":"created with sdata v{}.".format(__version__),
                 "now":"{}".format(now_utc_str()),
                 }

        tmpl = """<!DOCTYPE html>
<html lang="de">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
    h1 {{ 
       background-color: #00FFFF77;
       color: black;
    }}
    h2 {{ 
       background-color: #00FFFF44;
       color: black;
    }}
    h3 {{ 
       background-color: #00FFFF11;
       color: black;
    }}
    p {{
    	color: black;
    }}
    table, th, td, caption {{
    border: 1px solid #a0a0a0;
    }}
    
    table {{
      border-collapse: collapse;
      border-spacing: 1em;
      border-width: thin 0 0 thin;
      margin: 0 0 1em;
      table-layout: auto;
      max-width: 100%;
      text-align: right;
    }}
    th, td {{
      font-weight: normal;
      text-align: left;
      border-spacing: 1em;
      padding: .1em .3em;
    }}
    th, caption {{
      background-color: #f1f3f4;
      font-weight: 700;
    }}
    </style>
  </head>
  <body>
    <h1>{title}</h1>
    <h2>Download</h2>
    <p">{xlsx_tag}</p>
    <h2>Description</h2>
    <p>{description}</p>
    <h2>Metadata</h2>
    {metadata}
    <h2>Table</h2>
    {table}
    <h3>Table Description</h3>
    {table_description}
    <p>{sdata}</p>
    <p>{now}</p>
  </body>
</html>""".format(**param)
        try:
            with open(filepath, "w") as fh:
                fh.write(tmpl)
        except Exception as exp:
            raise

    def copy(self):
        """create a copy of the Data object

        .. code-block:: python

            data = sdata.Data(name="data", uuid="38b26864e7794f5182d38459bab85842", description="this is remarkable")
            datac = data.copy()
            print("data  {0.uuid}".format(data))
            print("datac {0.uuid}".format(datac))
            print("datac.metadata['!sdata_parent'] {0.value}".format(datac.metadata["sdata_parent"]))

        .. code-block::

            data  38b26864e7794f5182d38459bab85842
            datac 2c4eb15900af435d8cd9c8573ca777e2
            datac.metadata['!sdata_parent'] 38b26864e7794f5182d38459bab85842

        :return: Data
        """
        data = copy.deepcopy(self)
        data.metadata.add(self.SDATA_PARENT, self.uuid)
        data.metadata.add(self.SDATA_UUID, self.gen_uuid())
        data.metadata.add(self.SDATA_MTIME, now_utc_str(), dtype="str")
        return data

    def gen_uuid(self):
        """generate new uuid string

        :return: str, e.g. '5fa04a3738e4431dbc34eccea5e795c4'
        """
        return uuid.uuid4().hex

    def refactor(self, fix_columns=True, add_table_metadata=True):
        """helper function

        * to cleanup dataframe column name
        * to define Attributes for all dataframe columns
        """
        if isinstance(self.table, pd.DataFrame):
            mapper = {}
            for old_colname in self.table.columns:
                name, unit = extract_name_unit(old_colname)
                if fix_columns:
                    mapper[old_colname] = name
                if add_table_metadata:
                    old_attr = self.metadata.get(old_colname)
                    if old_attr:
                        logger.info("skip: {}".format(old_attr))
                        self.metadata.relabel(old_colname, name)
                    else:
                        self.metadata.add(name=name, description=old_colname, unit=unit, dtype="float")
            self.table.rename(columns=mapper, inplace=True)

SDATACLS = {"Data": Data,
            }
#
# __all__ = ["Data", Data]


def print_classes():
    for name, obj in inspect.getmembers(sys.modules[__name__]):
        if inspect.isclass(obj):
            print(obj)
